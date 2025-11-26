# -*- coding: utf-8 -*-
"""
Module d'export multi-formats pour publication
"""
from src.database import get_db, Article
from src.analytics import get_included_articles_summary
import csv
import json
from pathlib import Path
from datetime import datetime


def export_csv(output_path: str = None) -> str:
    """Export CSV des articles inclus"""
    
    db = next(get_db())
    articles = get_included_articles_summary(db)
    db.close()
    
    if not articles:
        print("⚠️ Aucun article inclus à exporter")
        return None
    
    if output_path is None:
        output_path = Path("data") / f"included_articles_{datetime.now().strftime('%Y%m%d')}.csv"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Colonnes
    fieldnames = ['id', 'title', 'authors', 'year', 'source', 'doi', 'link', 'reviewer', 'reviewed_at']
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(articles)
    
    print(f"✅ Export CSV : {len(articles)} articles → {output_path}")
    return str(output_path)


def export_json(output_path: str = None) -> str:
    """Export JSON complet"""
    
    db = next(get_db())
    articles = get_included_articles_summary(db)
    db.close()
    
    if not articles:
        print("⚠️ Aucun article inclus à exporter")
        return None
    
    if output_path is None:
        output_path = Path("data") / f"included_articles_{datetime.now().strftime('%Y%m%d')}.json"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    data = {
        'export_date': datetime.now().isoformat(),
        'total_articles': len(articles),
        'articles': articles
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Export JSON : {len(articles)} articles → {output_path}")
    return str(output_path)


def export_bibtex(output_path: str = None) -> str:
    """Export BibTeX pour citations"""
    
    db = next(get_db())
    included = db.query(Article).filter(Article.status == 'INCLUDED').all()
    db.close()
    
    if not included:
        print("⚠️ Aucun article inclus à exporter")
        return None
    
    if output_path is None:
        output_path = Path("data") / f"references_{datetime.now().strftime('%Y%m%d')}.bib"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for idx, article in enumerate(included, 1):
            # Générer clé BibTeX
            first_author = article.authors.split(',')[0].strip() if article.authors else "Unknown"
            first_author = first_author.split()[-1] if first_author else "Unknown"
            year = article.year or "XXXX"
            key = f"{first_author}{year}_{idx}"
            
            # Type d'entrée
            entry_type = "article"
            
            # Construire entrée BibTeX
            f.write(f"@{entry_type}{{{key},\n")
            f.write(f"  title = {{{article.title}}},\n")
            
            if article.authors:
                f.write(f"  author = {{{article.authors}}},\n")
            
            if article.year:
                f.write(f"  year = {{{article.year}}},\n")
            
            if article.doi:
                f.write(f"  doi = {{{article.doi}}},\n")
            
            if article.link:
                f.write(f"  url = {{{article.link}}},\n")
            
            f.write("}\n\n")
    
    print(f"✅ Export BibTeX : {len(included)} citations → {output_path}")
    return str(output_path)


def export_excel(output_path: str = None) -> str:
    """Export Excel (nécessite openpyxl)"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("⚠️ Module openpyxl non installé. Utiliser CSV à la place.")
        return export_csv(output_path.replace('.xlsx', '.csv') if output_path else None)
    
    db = next(get_db())
    articles = get_included_articles_summary(db)
    db.close()
    
    if not articles:
        print("⚠️ Aucun article inclus à exporter")
        return None
    
    if output_path is None:
        output_path = Path("data") / f"included_articles_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Créer workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles Inclus"
    
    # Header
    headers = ['ID', 'Titre', 'Auteurs', 'Année', 'Source', 'DOI', 'Lien', 'Reviewer', 'Date Revue']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for article in articles:
        ws.append([
            article['id'],
            article['title'],
            article['authors'],
            article['year'],
            article['source'],
            article['doi'],
            article['link'],
            article['reviewer'],
            article['reviewed_at']
        ])
    
    # Ajuster largeurs colonnes
    column_widths = [8, 50, 30, 8, 15, 25, 40, 15, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    wb.save(output_path)
    
    print(f"✅ Export Excel : {len(articles)} articles → {output_path}")
    return str(output_path)


if __name__ == "__main__":
    # Test
    print("Export CSV:")
    export_csv()
    
    print("\nExport JSON:")
    export_json()
    
    print("\nExport BibTeX:")
    export_bibtex()
